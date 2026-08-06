import re
import io
from collections import defaultdict

import streamlit as st
import openpyxl
from weasyprint import HTML

# --- CONFIGURATION DE LA PAGE ---
st.set_page_config(page_title="Générateur de Fiche UTS", layout="wide")

# Chemin vers le fichier Excel source. Peut aussi être uploadé via la sidebar.
DEFAULT_XLSX_PATH = "80_uts_test.xlsx"

# Dossier où se trouvent réellement les images (cartes, roues, profils) sur
# cette machine. Le fichier Excel contient des chemins Windows locaux
# (ex: C:\Users\...\9355.jpeg) qui n'existeront jamais tels quels ici :
# on ne garde que le nom de fichier (colonnes "*_simple") et on va le
# chercher dans ce dossier. Adaptez ce chemin à votre installation, ou
# laissez vide pour désactiver les images.
IMAGES_DIR = ""  # ex: "./images"


# ============================================================
# 1. CHARGEMENT ET INDEXATION DU FICHIER EXCEL
# ============================================================

@st.cache_data(show_spinner="Lecture du fichier Excel…")
def load_workbook_data(file_bytes: bytes):
    """
    Charge le classeur et construit :
    - header : liste des noms de colonnes (dans l'ordre du fichier)
    - columns_by_name : { "TAUX ARGILE_mod": [idx_strate1, idx_strate2, ...], ... }
      (les colonnes réelles sont préfixées "1 ", "2 ", ... dans le fichier ;
      on les regroupe par nom de base et on garde l'ordre = numéro de strate)
    - rows : toutes les lignes de données (une ligne = une UTS)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb["Feuil1"]

    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    data_rows = all_rows[1:]

    columns_by_name = defaultdict(list)
    for idx, h in enumerate(header):
        if not h or str(h).startswith("Colonne"):
            continue  # colonnes fantômes vides du fichier source
        h = str(h).strip()
        if h == "no_uts":
            columns_by_name["no_uts"].append(idx)
            continue
        m = re.match(r"^(\d+)\s+(.*)$", h)
        if m:
            name = m.group(2).strip()
            columns_by_name[name].append(idx)

    return header, columns_by_name, data_rows


def get_field(row, columns_by_name, name, occurrence=1, default=""):
    """Récupère la valeur d'un champ pour une strate donnée (1 = première strate)."""
    idxs = columns_by_name.get(name)
    if not idxs or occurrence > len(idxs):
        return default
    val = row[idxs[occurrence - 1]]
    if val is None or val == "":
        return default
    return val


def find_row_for_uts(rows, columns_by_name, id_uts):
    """Retourne la ligne correspondant au no_uts demandé, ou None."""
    idx_no_uts = columns_by_name["no_uts"][0]
    id_uts_str = str(id_uts).strip()
    for row in rows:
        if row[idx_no_uts] is not None and str(row[idx_no_uts]).strip() == id_uts_str:
            return row
    return None


def fmt_number(value, decimals=0):
    """Formate un nombre avec espace comme séparateur de milliers (style FR)."""
    if value in (None, "", "/"):
        return "/"
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    formatted = f"{num:,.{decimals}f}"
    formatted = formatted.replace(",", " ")
    if decimals == 0:
        formatted = formatted.split(".")[0]
    return formatted


def image_src(filename):
    """
    Construit une source d'image utilisable dans le HTML à partir du nom de
    fichier stocké dans le classeur (colonnes *_simple), en la cherchant dans
    IMAGES_DIR. Si le dossier n'est pas configuré ou le fichier introuvable,
    renvoie None (l'appelant doit alors ne pas afficher d'image).
    """
    if not filename or not IMAGES_DIR:
        return None
    import os
    path = os.path.join(IMAGES_DIR, filename)
    if os.path.isfile(path):
        return path
    return None


# ============================================================
# 2. CONSTRUCTION DES DONNÉES D'UNE FICHE À PARTIR D'UNE LIGNE
# ============================================================

STRATE_FIELDS = [
    ("rp_nom_strate", "Nom de la strate"),
    ("prof_appar_moy", "Profondeur (cm)"),
    ("ABONDANCE_EG_mod", "Abondance éléments grossiers (%)"),
    ("TEXTURE GEPPA", "Texture GEPPA"),
    ("TAUX ARGILE_mod", "Taux argile (g/kg)"),
    ("hydromorphie", "Hydromorphie"),
    ("EFFERVESCENCE", "Effervescence (classe)"),
    ("CALC_TOT_mod", "Calcaire total (g/kg)"),
    ("PH_EAU_mod", "pH eau"),
    ("TAUX MO_mod", "Taux MO (g/kg)"),
    ("CEC_mod", "CEC (cmol+/kg)"),
]


def build_fiche_data(row, columns_by_name):
    """Construit un dictionnaire propre avec toutes les valeurs utiles à la fiche."""
    def f(name, occurrence=1, default="/"):
        return get_field(row, columns_by_name, name, occurrence, default)

    data = {
        "id_uts": f("no_uts", default=""),
        "nom_uts": f("nom_uts"),
        "rp_2008_nom": f("rp_2008_nom"),
        "nom_ger": f("nom_ger"),
        "petite_region": f("petite_region"),
        "appellation_locale": f("appellation_locale"),
        "code_tariere": f("code_tariere"),
        "numero_typterre": f("UTT"),
        "typs": f("TypS"),
        "guide": f("guide"),
        "fiche": f("fiche"),
        "libelle_occupation": f("libelle_occupation"),
        "surf_uts": fmt_number(f("surf_uts", default=None)),
        "surf_ha": fmt_number(f("surf_ha", default=None)),
        "nom_mat": f("nom_mat"),
        "carte_img": image_src(f("chemin_carte_simple", default=None)),
        "profil_img": image_src(f("chemin_profil_simple", default=None)),
        "roue_agro_img": image_src(f("chemin_agro_simple", default=None)),
        "roue_services_img": image_src(f("chemin_services_simple", default=None)),
    }

    # Strates : on récupère les occurrences 1..6, on garde celles qui ont
    # au moins un nom de strate ou une profondeur renseignée.
    strates = []
    for occ in range(1, 7):
        nom = get_field(row, columns_by_name, "rp_nom_strate", occ, default=None)
        prof = get_field(row, columns_by_name, "prof_appar_moy", occ, default=None)
        if nom is None and prof is None:
            continue
        strate = {}
        for field_name, _label in STRATE_FIELDS:
            strate[field_name] = get_field(row, columns_by_name, field_name, occ, default="/")
        strates.append(strate)
    data["strates"] = strates

    return data


# ============================================================
# 3. GABARIT HTML
# ============================================================

HTML_HEAD_CSS = """
<style>
    @page { size: A4 landscape; margin: 0; }
    * { box-sizing: border-box; }
    body { margin: 0; padding: 0; background-color: #fff; font-family: Arial, sans-serif; }
    .fiche-layout { width: 1280px; height: 720px; padding: 15px 30px; background-color: #ffffff; position: relative; margin: 0 auto; }
    .fiche-header-container { width: 100%; margin-bottom: 15px; }
    .fiche-title-bar { background-color: #7b0d00; padding: 6px 15px; display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #fff; }
    .fiche-title-bar h1 { color: #ffffff; font-size: 16px; font-weight: bold; margin: 0; }
    .fiche-title-bar .fiche-number { color: #e6a8d7; font-size: 20px; font-weight: bold; }
    .fiche-subtitle-bar { background-color: #d26a00; padding: 4px 15px; width: 70%; border: 1px solid #000; margin-top: 2px; }
    .fiche-subtitle-bar h2 { color: #ffffff; font-size: 13px; font-weight: bold; margin: 0; }
    .fiche-top-grid { display: grid; grid-template-columns: 3.5fr 1.5fr 2fr 1.5fr; width: 100%; gap: 15px; margin-bottom: 15px; }
    .fiche-meta-info { font-size: 12px; line-height: 1.4; color: #000; }
    .fiche-meta-info div { display: flex; flex-wrap: wrap; column-gap: 15px; }
    .fiche-meta-info span { font-weight: bold; }
    .fiche-yellow-box { background-color: #fff8da; padding: 10px; font-size: 12px; font-weight: bold; height: max-content; color: #000; }
    .fiche-yellow-box p { margin: 0 0 5px 0; }
    .fiche-map { display: flex; flex-direction: column; align-items: flex-end; font-size: 10px; font-weight: bold; }
    .fiche-map img { width: 120px; height: auto; max-height: 90px; object-fit: contain; }
    .fiche-table-container { width: 100%; margin-bottom: 15px; }
    .fiche-table { width: 100%; border-collapse: collapse; font-size: 12px; color: #000; }
    .fiche-table th, .fiche-table td { border: 2px solid #000; padding: 6px; text-align: center; }
    .fiche-table th { font-weight: bold; background-color: #fff; }
    .fiche-table.grand-ensemble th, .fiche-table.grand-ensemble td { text-align: left; }
    .fiche-middle-section { display: flex; width: 100%; gap: 15px; margin-bottom: 10px; align-items: flex-start; }
    .fiche-profile { width: 200px; flex-shrink: 0; }
    .fiche-profile img { width: 100%; height: 180px; object-fit: cover; border: 1px solid #000; }
    .fiche-bottom-section { display: flex; width: 100%; justify-content: space-between; align-items: flex-start; border-top: 2px solid #8b5a2b; padding-top: 15px; margin-top: 5px; }
    .fiche-legend { display: flex; flex-direction: column; gap: 6px; font-size: 11px; color: #000; }
    .fiche-legend p { margin: 0 0 5px 0; }
    .fiche-legend-item { display: flex; align-items: center; gap: 10px; }
    .fiche-legend-color { width: 40px; height: 15px; border: 1px solid #000; }
    .fiche-wheels { display: flex; gap: 40px; justify-content: center; flex-grow: 1; align-items: flex-start; }
    .fiche-wheel-box { display: flex; flex-direction: column; align-items: center; width: 220px; }
    .fiche-wheel-title { font-size: 12px; font-weight: bold; text-align: center; margin-bottom: 10px; }
    .fiche-wheel-box img { width: 160px; height: 160px; object-fit: contain; }
    .fiche-logos { display: flex; flex-direction: column; align-items: flex-end; justify-content: flex-end; font-size: 9px; height: 100%; }
    .fiche-logos img { height: 30px; }
    .fiche-missing-img { font-size: 10px; color: #999; border: 1px dashed #ccc; padding: 20px; text-align: center; }
</style>
"""


def render_img_or_placeholder(src, css_class, alt):
    if src:
        return f'<img src="{src}" alt="{alt}">'
    return f'<div class="fiche-missing-img">{alt} non disponible</div>'


def build_strates_rows_html(strates):
    rows_html = ""
    for s in strates:
        rows_html += (
            "<tr>"
            f"<td><strong>{s['rp_nom_strate']}</strong></td>"
            f"<td>{s['prof_appar_moy']}</td>"
            f"<td>{s['ABONDANCE_EG_mod']}</td>"
            f"<td>{s['TEXTURE GEPPA']}</td>"
            f"<td>{s['TAUX ARGILE_mod']}</td>"
            f"<td>{s['hydromorphie']}</td>"
            f"<td>{s['EFFERVESCENCE']}</td>"
            f"<td>{s['CALC_TOT_mod']}</td>"
            f"<td>{s['PH_EAU_mod']}</td>"
            f"<td>{s['TAUX MO_mod']}</td>"
            f"<td>{s['CEC_mod']}</td>"
            "</tr>"
        )
    return rows_html


def build_html(data):
    strates_html = build_strates_rows_html(data["strates"])

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    {HTML_HEAD_CSS}
</head>
<body>
    <div class="fiche-layout">
        <div class="fiche-header-container">
            <div class="fiche-title-bar">
                <h1>{data['nom_uts']}</h1>
                <div class="fiche-number">Fiche n° {data['fiche']}</div>
            </div>
            <div class="fiche-subtitle-bar">
                <h2>{data['rp_2008_nom']}</h2>
            </div>
        </div>

        <div class="fiche-top-grid">
            <div class="fiche-meta-info">
                <div><span>Petite région naturelle :</span> {data['petite_region']} &nbsp;&nbsp;&nbsp; <span>Appellation locale :</span> {data['appellation_locale']}</div>
                <div><span>Code tarière :</span> {data['code_tariere']} &nbsp;&nbsp;&nbsp; <span>Numéro Typterre :</span> {data['numero_typterre']} &nbsp;&nbsp;&nbsp; <span>TypS :</span> {data['typs']} &nbsp;&nbsp;&nbsp; <span>Guide sol :</span> {data['guide']} Fiche: {data['fiche']}</div>
                <div><span>Occupation du sol :</span> {data['libelle_occupation']}</div>
            </div>

            <div class="fiche-yellow-box">
                <p>N°UTS : {data['id_uts']}</p>
                <p>Surface UTS : {data['surf_uts']} ha</p>
            </div>

            <div class="fiche-topo">
                <p>POSITION TOPOGRAPHIQUE</p>
                <p style="font-size: 10px; margin-top: 3px;">(non disponible dans le fichier source)</p>
            </div>

            <div class="fiche-map">
                <div>GER : {data['nom_ger']}<br>UTS : {data['id_uts']}</div>
                {render_img_or_placeholder(data['carte_img'], 'fiche-map', 'Carte')}
            </div>
        </div>

        <div class="fiche-table-container">
            <table class="fiche-table grand-ensemble">
                <thead><tr><th>Nom du Grand Ensemble de Référence du Géoportail</th><th>Matériau parental</th></tr></thead>
                <tbody><tr><td>{str(data['nom_ger']).upper()}<br>Surface {str(data['nom_ger']).upper()} en Grand Est : {data['surf_ha']} ha</td><td>{data['nom_mat']}</td></tr></tbody>
            </table>
        </div>

        <div class="fiche-middle-section">
            <div class="fiche-profile">
                {render_img_or_placeholder(data['profil_img'], 'fiche-profile', 'Profil de sol')}
            </div>
            <div class="fiche-table-container" style="margin-bottom: 0;">
                <table class="fiche-table">
                    <thead>
                        <tr>
                            <th>Nom de la<br>strate</th><th>Profondeur<br>(cm)</th><th>Abondance<br>éléments grossiers (%)</th>
                            <th>Texture<br>GEPPA</th><th>Taux argile<br>(g/kg)</th><th>Hydro-<br>morphie</th>
                            <th>Effervescence<br>(classe)</th><th>Calcaire<br>total (g/kg)</th><th>pH<br>eau</th>
                            <th>Taux MO<br>(g/kg)</th><th>CEC<br>(cmol+/kg)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {strates_html}
                    </tbody>
                </table>
            </div>
        </div>

        <div class="fiche-bottom-section">
            <div class="fiche-legend" style="width: 150px;">
                <p>Echelle des risques:</p>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #6a2c91;"></div>Risque fort</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #8871b6;"></div>Risque moyen</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #c3c0df;"></div>Risque faible</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #a0a0a0;"></div>Valeur non disponible</div>
            </div>

            <div class="fiche-wheels">
                <div class="fiche-wheel-box">
                    <div class="fiche-wheel-title">Roue agronomique<br>de l'UTS {data['id_uts']}</div>
                    {render_img_or_placeholder(data['roue_agro_img'], 'fiche-wheel-box', 'Roue agronomique')}
                </div>
                <div class="fiche-wheel-box">
                    <div class="fiche-wheel-title">Roue des services écosystémiques<br>de l'UTS {data['id_uts']}</div>
                    {render_img_or_placeholder(data['roue_services_img'], 'fiche-wheel-box', 'Roue des services')}
                </div>
            </div>

            <div class="fiche-legend" style="width: 170px;">
                <p>Echelle des potentialités:</p>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #1f7b3a;"></div>Valeur très élevée</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #82ca9c;"></div>Valeur élevée</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #b8e2b8;"></div>Valeur moyenne</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #ffffcc;"></div>Valeur faible</div>
                <div class="fiche-legend-item"><div class="fiche-legend-color" style="background-color: #a0a0a0;"></div>Valeur non disponible</div>
            </div>

            <div class="fiche-logos">
                <div style="display: flex; align-items: center; gap: 5px;">
                    <span style="font-size: 12px; font-weight: bold; color: #8b5a2b;">Gis Sol</span>
                    <span style="font-size: 10px; color: #006400; font-weight: bold; line-height: 1;">AGRICULTURES<br>&amp; TERRITOIRES</span>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""


# ============================================================
# 4. INTERFACE STREAMLIT
# ============================================================

st.title("📄 Générateur de Fiches UTS (Gis Sol)")

with st.sidebar:
    st.header("Source de données")
    uploaded_file = st.file_uploader("Fichier Excel des UTS (.xlsx)", type=["xlsx"])

    file_bytes = None
    if uploaded_file is not None:
        file_bytes = uploaded_file.read()
    else:
        try:
            with open(DEFAULT_XLSX_PATH, "rb") as f:
                file_bytes = f.read()
            st.caption(f"Fichier par défaut chargé : {DEFAULT_XLSX_PATH}")
        except FileNotFoundError:
            st.warning("Aucun fichier chargé. Importez un .xlsx ci-dessus.")

    st.markdown("---")
    st.header("Paramètres")

    if file_bytes:
        header, columns_by_name, data_rows = load_workbook_data(file_bytes)
        idx_no_uts = columns_by_name["no_uts"][0]
        available_ids = [str(r[idx_no_uts]) for r in data_rows if r[idx_no_uts] is not None]
        id_uts = st.selectbox("Choisissez l'UTS :", options=available_ids) if available_ids else None
    else:
        header = columns_by_name = data_rows = None
        id_uts = st.text_input("Saisissez l'ID UTS :", value="")

    st.markdown("---")
    st.subheader("Export")
    generate = st.button("Générer le PDF 📥", use_container_width=True)


# 2. Zone principale d'affichage
if not file_bytes:
    st.info("Chargez un fichier Excel pour commencer.")
    st.stop()

row = find_row_for_uts(data_rows, columns_by_name, id_uts) if id_uts else None

if row is None:
    st.error(f"Aucune UTS trouvée pour l'ID « {id_uts} ». Vérifiez la valeur saisie.")
    st.stop()

fiche_data = build_fiche_data(row, columns_by_name)
final_html = build_html(fiche_data)

st.subheader("Aperçu de la fiche")
st.markdown(f"Prévisualisation en direct pour l'UTS : **{id_uts}**")
st.components.v1.html(final_html, height=750, scrolling=True)

if generate:
    with st.spinner("Génération du PDF en cours..."):
        try:
            # base_url permet à WeasyPrint de résoudre les chemins d'images
            # relatifs (ceux construits depuis IMAGES_DIR) comme des fichiers
            # locaux plutôt que de les interpréter comme des URL.
            pdf_bytes = HTML(string=final_html, base_url=".").write_pdf()

            st.success("PDF généré avec succès !")
            st.download_button(
                label="Télécharger le fichier PDF",
                data=pdf_bytes,
                file_name=f"Fiche_UTS_{id_uts}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception as e:
            st.error("Erreur lors de la génération du PDF.")
            st.exception(e)